import os
from django.conf import settings
from django.db.models.functions import ExtractHour

from django.shortcuts import redirect, render
from django.urls import reverse
from db.models import *
from django.http import JsonResponse
from django.db.models import Sum, Avg, Count
from datetime import datetime, timedelta

from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

names = []
time_ranges = [(10, 11), (11, 12), (13, 14), (14, 15), (15, 16), (16, 17), (17, 18)]

for u in T_USERS.objects.all():
    names.append(u.username)

company = ['naver', 'navergfa', 'gmarket', 'eleven', 'wemakeprice']

change = ['네이버', '네이버지에프에이', '지마켓', '11번가', '위메프']

data = {}

for n in names:
    data[n] = {}
    
  ###########
 # 에러처리 #
###########
def e404(request):
    return render(request, 'search/e404.html',)

  ##########################
 #전주의 시작일, 끝일 구하기#
##########################
def get_previous_week_dates(reference_date):
    start_of_week = reference_date - timedelta(days=reference_date.weekday())  # 주의 시작 (월요일)
    end_of_week = start_of_week + timedelta(days=6)  # 주의 끝 (일요일)
    return start_of_week, end_of_week

  ##########################
 #전월의 시작일, 말일 구하기#
##########################
def get_previous_month_range(current_date):
    # 전월의 마지막 날을 구하기 위해 현재 월의 첫날에서 하루를 빼면 전월의 마지막 날이 됩니다.
    end_of_last_month = current_date.replace(day=1) - timedelta(days=1)
    # 전월의 첫날은 마지막 날의 day를 1로 설정하면 됩니다.
    start_of_last_month = end_of_last_month.replace(day=1)
    return start_of_last_month, end_of_last_month

  ##############################
 #전전주 대비 전주 성장률 구하기#
##############################
def calculate_growth(last_week_total, week_before_last_total):
    if week_before_last_total == 0 and last_week_total != 0:
        return 100
    elif week_before_last_total == 0 and last_week_total == 0:
        return 0
    else:
        return (last_week_total / week_before_last_total * 100)

  #########################################################
 # 초를 시 분 초 형식으로 00:00:00 로 표시하도록 만드는 함수 #
##########################################################
def seconds_to_hms(seconds):
    # 반올림 처리
    seconds = round(seconds)
    # 시간, 분, 초 계산
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60
    # 포매팅된 문자열 반환
    return f"{hours:02}:{minutes:02}:{seconds:02}"

  ###############################################################
 # 전주, 전전주 평균콜수, 평균콜시간, 전주 시간대별 주간 평균통화량 #
###############################################################
def calculate_weekly_call_data(data, today, user_data, names):
    last_week_start, last_week_end = get_previous_week_dates(today - timedelta(weeks=1))
    week_before_last_start, week_before_last_end = get_previous_week_dates(today - timedelta(weeks=2))
    
    # 전주 평균 콜 시간과 콜 수
    last_week_avg_data = T_CALL_DAY_LIST.objects.filter(
        call_date__range=[last_week_start, last_week_end],
        sender__in = user_data
    ).values('sender').annotate(
        avg_call_duration=Avg('time_to_sec'),
        call_count=Count('call_id')
    )

    # 전전주 평균 콜 시간과 콜 수
    week_before_last_avg_data = T_CALL_DAY_LIST.objects.filter(
        call_date__range=[week_before_last_start, week_before_last_end],
        sender__in = user_data
    ).values('sender').annotate(
        avg_call_duration=Avg('time_to_sec'),
        call_count=Count('call_id')
    )

    for call in last_week_avg_data:
        data[user_data[call['sender']]]['last_avg_call_duration'] = seconds_to_hms(call['avg_call_duration']) if call['avg_call_duration'] else seconds_to_hms(0)
        data[user_data[call['sender']]]['last_call_count'] = round(call['call_count']/7,1)
    for call in week_before_last_avg_data:
        data[user_data[call['sender']]]['b_last_avg_call_duration'] = seconds_to_hms(call['avg_call_duration']) if call['avg_call_duration'] else seconds_to_hms(0)
        data[user_data[call['sender']]]['b_last_call_count'] = round(call['call_count']/7,1)
    
    # 전주 시간대별 주간 평균 통화량 계산
    base_query = T_CALL_DAY_LIST.objects.filter(
        call_date__range=[last_week_start, last_week_end],
        sender__in=user_data.keys()
    ).annotate(
        hour=ExtractHour('call_date')  # 시간 추출
    )
    
    for sender in user_data.keys():
        sender_data = base_query.filter(sender=sender)
        for start, end in time_ranges:
            avg_data = sender_data.filter(
                hour__gte=start,
                hour__lt=end
            ).aggregate(
                avg_call_duration=Avg('call_duration')
            )
            
            avg_duration = avg_data['avg_call_duration']
            if avg_duration is None:
                avg_duration = "00:00:00"  # 적절한 기본값 설정
            else:
                avg_duration = seconds_to_hms(avg_duration)

            # 결과 저장
            data[user_data[sender]][f'_{start}{end}avg_duration'] = avg_duration
    return data

  #######################
 # 엑셀생성시 기본설정값 #
#######################
header_font = Font(bold=True)
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center_align = Alignment(horizontal='center', vertical='center')

  ##################
 # 첫번째 헤더생성 #
##################
def create_excel_header(worksheet, company_list):
    
      #################
     # 메인 글씨 작성 #
    ################# 
    cell = worksheet.cell(row=1, column=1)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=21)
    cell.value = '주간 업무 미팅 자료'
    cell.font = Font(name='Calibri', size=16, bold=True)
    cell.alignment = header_align
    cell.border = border
    
      ########################
     # 첫 번째 행의 헤더 설정 #
    ########################
    worksheet['A4'] = '매체'
    worksheet['A4'].font = header_font
    worksheet['A4'].fill = header_fill
    worksheet['A4'].alignment = header_align
    worksheet['A4'].border = border
    col_idx = 2  # B열에서 시작
    
    # 매체 설정
    for company in change:
        cell = worksheet.cell(row=4, column=col_idx)
        worksheet.merge_cells(start_row=4, start_column=col_idx, end_row=4, end_column=col_idx+2)
        cell.value = company
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        col_idx += 3
        
    # '맵핑수(누적)'과 '전매체 Live 계정수' 헤더 설정
    cell = worksheet.cell(row=4, column=col_idx)
    worksheet.merge_cells(start_row=4, start_column=col_idx, end_row=4, end_column=col_idx+1)
    cell.value = '맵핑수(누적)'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    col_idx += 2

    cell = worksheet.cell(row=4, column=col_idx)
    worksheet.merge_cells(start_row=4, start_column=col_idx, end_row=4, end_column=col_idx+2)
    cell.value = '전매체 Live 계정수'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    col_idx += 3

    # 두 번째 행의 헤더 설정
    worksheet['A5'] = '팀원'
    worksheet['A5'].font = header_font
    worksheet['A5'].fill = header_fill
    worksheet['A5'].alignment = header_align
    worksheet['A5'].border = border

    col_idx = 2  # B열에서 시작
    
    # 회사별 전전주매출, 전주매출, 성장률 설정
    sub_headers = ['전전주매출', '전주매출', '성장률']
    for _ in company_list:
        for sub_header in sub_headers:
            cell = worksheet.cell(row=5, column=col_idx)
            cell.value = sub_header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            col_idx += 1

    
    # '맵핑수(누적)' 밑에 '신규', '이관' 설정
    cell = worksheet.cell(row=5, column=col_idx)
    cell.value = '신규'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    col_idx += 1
    
    cell = worksheet.cell(row=5, column=col_idx)
    cell.value = '이관'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    col_idx += 1


    # '전매체 Live 계정수' 밑에 '전전주', '전주', '증감' 설정
    sub_headers = ['전전주', '전주', '증감']
    for sub_header in sub_headers:
        cell = worksheet.cell(row=5, column=col_idx)
        cell.value = sub_header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        col_idx += 1

  ##################
 # 두번째 헤더생성 #
##################
def create_excel_header2(worksheet, row_num):
    col_idx = 1
    
    cell = worksheet.cell(row=row_num, column=col_idx)
    worksheet.merge_cells(start_row=row_num, start_column=col_idx, end_row=row_num+1, end_column=col_idx)
    cell.value = '팀원'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    col_idx += 1
    
    headers = ['업무', '평균콜수', '평균콜시간', '평균콜수', '평균콜시간']
    # 첫 번째 행의 헤더 설정
    for sub_header in headers:
        cell = worksheet.cell(row=row_num, column=col_idx)
        cell.value = sub_header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        col_idx += 1
    
    cell = worksheet.cell(row=row_num, column=col_idx)
    worksheet.merge_cells(start_row=row_num, start_column=col_idx, end_row=row_num+1, end_column=col_idx)
    cell.value = '팀원'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    col_idx += 1
    
    cell = worksheet.cell(row=row_num, column=col_idx)
    worksheet.merge_cells(start_row=row_num, start_column=col_idx, end_row=row_num, end_column=col_idx+6)
    cell.value = '시간대별 주간 평균통화량'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    
    # 두번째 행
    col_idx = 2
    row_num+=1

    cell = worksheet.cell(row=row_num, column=col_idx)
    cell.value = '집중시간'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    col_idx += 1
    
    cell = worksheet.cell(row=row_num, column=col_idx)
    worksheet.merge_cells(start_row=row_num, start_column=col_idx, end_row=row_num, end_column=col_idx+1)
    cell.value = '전전주'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    col_idx+=2
    
    cell = worksheet.cell(row=row_num, column=col_idx)
    worksheet.merge_cells(start_row=row_num, start_column=col_idx, end_row=row_num, end_column=col_idx+1)
    cell.value = '전주'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    col_idx+=3
    
    headers = ['10시 ~ 11시','11시 ~ 12시','1시 ~ 2시','2시 ~ 3시','3시 ~ 4시','4시 ~ 5시','5시 ~ 6시']
    # 첫 번째 행의 헤더 설정
    for sub_header in headers:
        cell = worksheet.cell(row=row_num, column=col_idx)
        cell.value = sub_header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        col_idx += 1
    
    return row_num


  ##################
 # 세번째 헤더생성 #
##################
def create_excel_header3(worksheet, row_num):
    col_idx = 1
    
    cell = worksheet.cell(row=row_num, column=col_idx)
    worksheet.merge_cells(start_row=row_num, start_column=col_idx, end_row=row_num+1, end_column=col_idx)
    cell.value = '팀원'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    col_idx += 1
    
    # 첫 번째 행의 헤더 설정
    cell = worksheet.cell(row=row_num, column=col_idx)
    worksheet.merge_cells(start_row=row_num, start_column=col_idx, end_row=row_num, end_column=col_idx+3)
    cell.value = '전매체 총 매출'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    col_idx += 4
    
    cell = worksheet.cell(row=row_num, column=col_idx)
    worksheet.merge_cells(start_row=row_num, start_column=col_idx, end_row=row_num+1, end_column=col_idx)
    cell.value = '주력매체'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    col_idx += 1
    
    cell = worksheet.cell(row=row_num, column=col_idx)
    worksheet.merge_cells(start_row=row_num, start_column=col_idx, end_row=row_num+1, end_column=col_idx)
    cell.value = 'DB수집방식'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    col_idx += 1

    cell = worksheet.cell(row=row_num, column=col_idx)
    worksheet.merge_cells(start_row=row_num, start_column=col_idx, end_row=row_num+1, end_column=col_idx)
    cell.value = '영업방식'
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
    col_idx += 1
    
    # 두번째 행
    col_idx = 2
    row_num+=1
    
    sub_headers = ['전월매출', '당월누적매출', '예상매출', '예상증감']
    for sub_header in sub_headers:
        cell = worksheet.cell(row=row_num, column=col_idx)
        cell.value = sub_header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        col_idx += 1
    return row_num

  ##########################
 # 데이터를 엑셀로 출력하기 #
##########################
def export_to_excel(data, selected_names, company_list):
    # 엑셀 파일 생성을 위한 버퍼
    wb = Workbook()
    ws = wb.active
    ws.title = "주간 업무 미팅 자료"
    today = datetime.now().date()
    filename = str(today)
    
    # 첫번째 테이블 헤더 생성
    create_excel_header(ws, company_list)
    
    # 첫번째 테이블 데이터 쓰기
    row_num = 6
    for name in selected_names:
        filename += name + '_'
        if name in data:
            cell = ws.cell(row=row_num, column=1, value=name)
            cell.alignment = center_align
            col_idx = 2
            for company in company_list:
                company_data = data[name].get(company, {})
                
                cell = ws.cell(row=row_num, column=col_idx, value=company_data.get('week_before_last_total', 0))
                cell.alignment = center_align
                
                cell = ws.cell(row=row_num, column=col_idx+1, value=company_data.get('last_week_total', 0))
                cell.alignment = center_align
                
                cell = ws.cell(row=row_num, column=col_idx+2, value=f"{company_data.get('growth_rate', 0)}%")
                cell.alignment = center_align
                
                col_idx += 3
    
            cell = ws.cell(row=row_num, column=col_idx, value=data[name].get('new'))
            cell.alignment = center_align
            
            cell = ws.cell(row=row_num, column=col_idx+1, value=data[name].get('escalation'))
            cell.alignment = center_align
            
            cell = ws.cell(row=row_num, column=col_idx+2, value=data[name].get('before_last_live'))
            cell.alignment = center_align
            
            cell = ws.cell(row=row_num, column=col_idx+3, value=data[name].get('last_live'))
            cell.alignment = center_align
            
            cell = ws.cell(row=row_num, column=col_idx+4, value=data[name].get('increase'))
            cell.alignment = center_align
        row_num += 1
    
    # 두번째 테이블 헤더 생성
    row_num = create_excel_header2(ws, row_num+1)
    row_num+=1
    
    # 두번째 테이블 데이터 작성
    for name in selected_names:
        cell = ws.cell(row=row_num, column=1, value=name)
        cell.alignment = center_align
        
        col_idx = 2
        
        cell = ws.cell(row=row_num, column=col_idx, value=data[name].get('focus'))
        cell.alignment = center_align
        
        cell = ws.cell(row=row_num, column=col_idx+1, value=data[name].get('b_last_call_count', 0))
        cell.alignment = center_align
        
        cell = ws.cell(row=row_num, column=col_idx+2, value=data[name].get('b_last_avg_call_duration', '00:00:00'))
        cell.alignment = center_align
        
        cell = ws.cell(row=row_num, column=col_idx+3, value=data[name].get('last_call_count', 0))
        cell.alignment = center_align
        
        cell = ws.cell(row=row_num, column=col_idx+4, value=data[name].get('last_avg_call_duration', '00:00:00'))
        cell.alignment = center_align
        
        cell = ws.cell(row=row_num, column=col_idx+5, value=name)
        cell.alignment = center_align
        
        for start,end in time_ranges:
            if data[name].get(f'_{start}{end}avg_duration') == None:
                res = '00:00:00'
            else:
                res = data[name].get(f'_{start}{end}avg_duration')
            cell = ws.cell(row=row_num, column=col_idx+6, value=res)
            cell.alignment = center_align
            col_idx +=1
        
        row_num += 1
    
    # 세번째 테이블 헤더 생성
    row_num = create_excel_header3(ws, row_num+1)
    row_num+=1
    
    # 세번째 테이블 데이터 작성
    for name in selected_names:
        
        if name in data:
            # 각 팀원에 대한 데이터를 작성하기 위해 새로운 행을 추가합니다.
            cell = ws.cell(row=row_num, column=1, value=name)
            cell.alignment = center_align
            
            col_idx = 2
            
            cell = ws.cell(row=row_num, column=col_idx, value=data[name].get('previous_month_sales', 0))
            cell.alignment = center_align
            
            cell = ws.cell(row=row_num, column=col_idx+1, value=data[name].get('current_month_sales', 0))
            cell.alignment = center_align
            
            cell = ws.cell(row=row_num, column=col_idx+2, value=data[name].get('estimated_sales', 0))
            cell.alignment = center_align
            
            cell = ws.cell(row=row_num, column=col_idx+3, value=data[name].get('estimated_growth', 0))
            cell.alignment = center_align
            

            col_idx += 4
            row_num += 1
    
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column if cell.value]  # 값이 있는 셀만 고려
        for cell in column:
            try:
                # 셀의 내용 길이를 체크 (셀에 저장된 데이터가 문자열이 아닐 수도 있으니 str()로 변환)
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except:
                pass
        adjusted_width = (max_length + 4)  # 약간의 여백 추가
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    # 버퍼를 통해 파일 저장
    filename += ".xlsx"
    file_path = os.path.join(settings.MEDIA_ROOT, filename)
    wb.save(file_path)

    # 서버 내 파일 URL 생성
    file_url = settings.MEDIA_URL + filename
    return file_url

  ####################
 # views.py 메인함수 #
####################
def sales_report(request):
    
    # mpid 로그인 아이디 가져와서 세션저장
    mpid = request.GET.get('mpid')  # URL에서 mpid를 가져옵니다.
    session = request.session.get('mpid')

    # 관리자인지 아닌지 판단
    is_manager = session in {'mp001', 'mp003', 'mp008', 'mp8826', 'mp777', 'ceo', 'mp015', 'mp027', 'mp010'}
    xgroup =  ["경영지원팀", "기술지원팀", "퇴사자", "외부", "대대행"]
    
    # mpid 에따른 분기처리 및 url 숨기기
    if mpid:
        if session:
            if mpid != session:
                return redirect('e404')
        else:
            request.session['mpid'] = mpid
            return redirect(reverse('sales_report'))
    else:
        if not session:
            return redirect('e404')
        
    if is_manager:
        teams = T_DEPTS.objects.all()  # 관리자는 모든 팀 정보를 볼 수 있습니다.
        members = T_USERS.objects.all()  # 관리자는 모든 멤버 정보를 볼 수 있습니다.
        select = False
        user_dept_code = None

    else:
        dept_code = T_USERS.objects.get(userid=session).dept_code
        teams = T_DEPTS.objects.filter(dept_code=dept_code)  # 팀장은 자신의 팀만 조회합니다.
        members = T_USERS.objects.filter(dept_code__in=teams)  # 해당 팀 코드에 속하는 사용자들만 조회합니다.
        select = True
        user_dept_code = dept_code 

    return render(request, 'search/search.html', {
        'teams': teams,
        'members': members,
        'select' : select,
        'user_dept_code': user_dept_code,
        'xgroup' : xgroup,
    })
    
  #######################
 # 비동기 처리 메인함수  #
########################
def fetch_team_data(request):
    
    # 팀, 멤버 선택시 값 가져오기
    team_code = request.GET.get('team_code')
    member_code = request.GET.get('member_code')
    ref_date_str = request.GET.get('ref_date')
    export_excel = request.GET.get('export_excel') == 'true'
    
    # 문자열을 datetime.date 객체로 변환
    if ref_date_str:
        ref_date = datetime.strptime(ref_date_str, '%Y-%m-%d').date()
        today = ref_date
    else:
        today = datetime.now().date()

    # 팀 선택, 멤버 선택에 따른 조회할 유저 담기
    names = []
    company_list = {}
    
    if team_code and member_code:
        members = T_USERS.objects.filter(username=member_code)
    elif team_code:
        members = T_USERS.objects.filter(dept_code=team_code)
    elif member_code:
        members = T_USERS.objects.filter(username=member_code)
    else:
        members = T_USERS.objects.all()
    for m in members:
        names.append(m.username)
    
    for com in company:
        company_list[com] = 0
    
    # 조회할 유저 데이터 처리
    data = data_process(names,today)
    
    # T_Call_Day 의 sender 값에 해당하는 유저 이름을 찾기위해 user_data 에 key, value 값에 uphone, username 으로 저장
    user_data = {}
    for user in names:
        res = T_USERS.objects.get(username=user)
        user_data[res.uphone] = res.username
        
    # 전주, 전전주 평균콜수, 평균콜시간, 전주 시간대별 주간 평균통화량
    data = calculate_weekly_call_data(data, today, user_data, names)
    
    if export_excel:
        file_url = export_to_excel(data, names, company)
        full_url = request.build_absolute_uri(file_url)  # 완전한 URL 생성
        return JsonResponse({'url': full_url})
    return JsonResponse({'members': data, 'selected' : names, 'company': company_list}, safe=False)
  
  ##################
 # 데이터 처리하기 #
##################
def data_process(members, today):
    
    # 선택된 이름을 받아옵니다.
    data = {}
    
    # 현재 날짜를 기준으로 전주와 전전주 날짜를 계산합니다.
    last_week_start, last_week_end = get_previous_week_dates(today - timedelta(weeks=1))
    week_before_last_start, week_before_last_end = get_previous_week_dates(today - timedelta(weeks=2))

    # 현재 날짜를 기준으로 전월달 과 현재까지를 계산합니다.
    previous_month_start, previous_month_end = get_previous_month_range(today)
    current_month_start = today.replace(day=1)

    # 모든 관련 데이터를 한 번에 조회합니다.
    last_week_data = T_Sales_Day.objects.filter(
        media_id__in=company,
        mkt_nm__in=members,
        sale_date__range=[last_week_start, last_week_end]
    ).values('media_id', 'mkt_nm').annotate(total=Sum('tot_amt'))

    week_before_last_data = T_Sales_Day.objects.filter(
        media_id__in=company,
        mkt_nm__in=members,
        sale_date__range=[week_before_last_start, week_before_last_end]
    ).values('media_id', 'mkt_nm').annotate(total=Sum('tot_amt'))

    # 데이터를 파싱하여 처리합니다.
    last_week_totals = { (d['media_id'], d['mkt_nm']): d['total'] for d in last_week_data }

    week_before_last_totals = { (d['media_id'], d['mkt_nm']): d['total'] for d in week_before_last_data }

    # 지난 달과 현재까지의 데이터를 조회합니다.
    last_month_data = T_NETSALES.objects.filter(
        sale_month__gte=previous_month_start.strftime('%Y%m'),
        sale_month__lte=previous_month_end.strftime('%Y%m'),
        mkt_name__in=members
    ).values('mkt_name').annotate(total=Sum('tot_amt'))

    # 각 이름별로 데이터 처리
    for name in members:
        data[name] = {}

        # 전월 매출에 따른 집중 시간 설정
        last_month_total = next((item['total'] for item in last_month_data if item['mkt_name'] == name), 0)
        if 1000000 <= last_month_total < 4000000:
            focus = '2시간'
        elif 4000000 <= last_month_total < 8000000:
            focus = '1시간'
        elif 8000000 <= last_month_total < 10000000:
            focus = '30분'
        else:
            focus = '0분'
        
        data[name]['focus'] = focus
 
          #########################
         # 신규, 이관을 계산합니다 #
        #########################
        
        # 신규
        new = T_TRANSFER.objects.filter(
            mkt_nm=name,
            trns_gbn=1
        ).count()
        
        # 이관
        escalation = T_TRANSFER.objects.filter(
            mkt_nm=name,
            trns_gbn = 2
        ).count()
        
        data[name]['new'] = new
        data[name]['escalation'] = escalation    
            
          #################################
         # 전매체 Live 계정수를 계산합니다 #
        #################################   

        # 전전주 Live 계정수
        before_last_live = T_Sales_Day.objects.filter(
            mkt_nm=name,
            tot_amt__gt=0,  # tot_amt가 0보다 큰 조건 추가
            sale_date__range=[week_before_last_start, week_before_last_end]
        ).count()  # 객체의 개수를 세는 메소드 사용

        # 전주 Live 계정수
        last_live = T_Sales_Day.objects.filter(
            mkt_nm=name,
            tot_amt__gt=0,  # tot_amt가 0보다 큰 조건 추가
            sale_date__range=[last_week_start, last_week_end]
        ).count()  # 객체의 개수를 세는 메소드 사용
        
        # 증감
        increase = last_live-before_last_live
        
        data[name]['before_last_live'] = before_last_live
        data[name]['last_live'] = last_live
        data[name]['increase'] = increase

        
          ###########################################
         # 전월매출, 당월누적매출, 예상매출, 예상증감 #
        ###########################################
        
        # 전월 매출 
        previous_month_sales = round(T_Sales_Day.objects.filter(
                mkt_nm=name,
            sale_date__range=[previous_month_start, previous_month_end]
            ).aggregate(total=Sum('tot_amt'))['total'] or 0)

        # 당월 누적 매출
        current_month_sales = round(T_Sales_Day.objects.filter(
            mkt_nm=name,
            sale_date__range=[current_month_start, today]
        ).aggregate(total=Sum('tot_amt'))['total'] or 0)

        # 예상 매출 계산 (단순화된 예상치 계산)
        estimated_sales = round((current_month_sales / today.day) * 30 if today.day != 0 else 0
)
        # 예상 증감
        estimated_growth = round(estimated_sales - previous_month_sales)
        
        data[name]['previous_month_sales'] = previous_month_sales
        data[name]['current_month_sales'] = current_month_sales
        data[name]['estimated_sales'] = estimated_sales
        data[name]['estimated_growth'] = estimated_growth
        
        for com in company:
            last_week_total = last_week_totals.get((com, name), 0)
            week_before_last_total = week_before_last_totals.get((com, name), 0)
            growth_rate = calculate_growth(last_week_total, week_before_last_total)
            # 결과 저장
            data[name][com] = {
                'last_week_total': round(last_week_total),
                'week_before_last_total': round(week_before_last_total),
                'growth_rate': round(growth_rate)
            }
        
    return data
